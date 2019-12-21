import os
import openpyxl
import string
import webbrowser

def reverse(s):
    rev=""
    for i in range(len(s)-1,-1,-1):
        rev=rev+s[i]
    return rev

def arrdefine():
    letters=list(string.ascii_letters)
    digits=list(string.digits)
    l=letters+digits
    return l

l=arrdefine()

def urlshort(i):
    s=""
    while(i != 0):
        s=s+l[i%62]
        i=int(i/62)
    s=reverse(s)
    return s

curr_dir=os.path.dirname(__file__)
filename=os.path.join(curr_dir,"URLshortner.xlsx")

if os.path.isfile(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb['URLshort']
else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title="URLshort"
    sheet['A1']="ID"
    sheet['B1']="Long URL"
    sheet['C1']="Shorten URL"
    sheet['A2']="100000"
    wb.save(filename)

c=0
while(c!= 3):
    print("\n1. Create Short URL\n2. Try Short URL\n3. Exit")
    c=int(input("Enter your choice : "))
    if(c==1):
        url=input("Enter a URL : ")
        r=sheet.max_row
        val=sheet.cell(row=r, column=1).value
        surl = urlshort(int(val))
        sheet.cell(row=r,column=2,value=url)
        sheet.cell(row=r,column=3,value=surl)
        sheet.cell(row=r+1,column=1,value=int(val)+100)
        wb.save(filename)
        print("Short URL for",url,"is",surl)
    elif(c==2):
        url=input("Enter Short URL : ")
        activate=0
        for i in range(1,sheet.max_row):
            if(sheet.cell(row=i, column=3).value == url):
                url = sheet.cell(row=i, column=2).value
                activate=1
                break
        if(activate==0):
            print("URL not Found.")
        else:
            webbrowser.open(url)
    elif(c==3):
        pass
    else:
        print("Enter Correct Choice.")

    
    
