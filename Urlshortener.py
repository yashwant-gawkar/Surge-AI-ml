import os
import openpyxl
import string
import webbrowser

def reverse(s1):
    sr = ""
    for i in range(len(s1)-1,-1,-1):
        sr = sr + s1[i]
    return sr

def arrdefine():
    letters = list(string.ascii_letters)
    digits = list(string.digits)
    l = letters+digits
    return l

l = arrdefine()

def urlshort(i):
    s = ""
    while(i != 0):
        s = s + l[i%62]
        i = int(i/62)
    s = reverse(s)
    return s

curr_dir = os.path.dirname(__file__)
filename = os.path.join(curr_dir,"URLshortner.xlsx")

if os.path.isfile(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb['URLshort']
else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "URLshort"
    sheet['A1'] = "ID"
    sheet['B1'] = "Long URL"
    sheet['C1'] = "Short URL"
    sheet['A2'] = "100000"
    wb.save(filename)

ch = 0
while(ch != 3):
    print("\n1. Create Short URL\n2. Try Short URL\n3. Exit")
    ch = int(input("Enter your choice : "))
    if(ch == 1):
        url = input("Enter URL : ")
        mr = sheet.max_row
        val = sheet.cell(row=mr, column=1).value
        surl = urlshort(int(val))
        sheet.cell(row=mr, column=2, value=url)
        sheet.cell(row=mr, column=3, value=surl)
        sheet.cell(row=mr+1, column=1, value=int(val)+100)
        wb.save(filename)
        print("Short URL for",url,"is",surl)
    elif(ch == 2):
        url = input("Enter Short URL : ")
        activate = 0
        for i in range(1,sheet.max_row):
            if(sheet.cell(row=i, column=3).value == url):
                url = sheet.cell(row=i, column=2).value
                activate = 1
                break
        if(activate == 0):
            print("No URL Found. Type Correct URL.")
        else:
            webbrowser.open(url)
    elif(ch == 3):
        pass
    else:
        print("Enter Correct Choice.")

    
    
